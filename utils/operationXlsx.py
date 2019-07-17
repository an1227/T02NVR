#!/usr/bin/env python
# -*- coding=utf-8 -*-

"""
@作者：wuchengan
@文件名：T02NVR/operationXlsx.py
@时间：2019/7/8 11:16
@文档说明:
"""
import openpyxl, os, xlrd
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color, RED, GREEN

class OperationExcel(object):
    def dir_base(self, filename, filePath='data'):
        """
        获取data文件夹下的文件
        :param filename: 要读取的文件的名称
        :param filePath: 要读取的文件对应的文件夹
        :return:
        """
        return os.path.join(os.path.dirname(os.path.dirname(__file__)), filePath, filename)

    def getExcelData(self, filename):
        """
        读取Excel数据
        :param filename:Excel文件名
        :return: rows
        """
        rows = []
        with xlrd.open_workbook(self.dir_base(filename)) as f:   #用打开xls的方式open_workbook打开工作表
            sheet = f.sheet_by_index(0)#通过sheet索引获取Excel第一张表
            for row_index in range(1, sheet.nrows):
                rows.append(sheet.row_values(row_index, 0, sheet.ncols-1))#row_values(行号，列号开始，列号结束)
            return rows

    def modifyExcel(self, filename, row, col, result, reason):
        """
        修改xlsx的单元格数据
        :param filename: 文件名
        :param row: 要修改的单元格行号
        :param col: 要修改的单元格列号
        :param result: 写入单元格的值
        :param reason: 备注
        :return:
        """
        work_book = openpyxl.load_workbook(self.dir_base(filename))  #用打开xlsx的方式load_workbook打开工作表
        work_sheet = work_book.worksheets[0]  #获取工作表达sheet页
        work_sheet.cell(row, col).value = result  # 写入执行结果
        work_sheet.cell(row, col).font = Font(bold=True)  # 执行结果字体加粗显示
        work_sheet.cell(row, col+1).value = reason  #写入执行备注
        if result == 'P':
            work_sheet.cell(row, col).fill = PatternFill('solid', fgColor=GREEN) #单元格用绿色填充，solid意思是纯色填充
        if result == 'F':
            work_sheet.cell(row, col).fill = PatternFill('solid', fgColor=RED)
        work_book.save(self.dir_base(filename))
# A=OperationExcel()
# print(A.getExcelData('data.xlsx'))
# # A.modifyExcel('data.xlsx', 1, 6, 'F', 'ss')
# # print(A.getExcelData('data.xlsx'))